"""Excel file handling module"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from typing import Dict, List, Optional
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelHandler:
    """Handle reading/writing to Excel expense tracking files"""

    def __init__(self, excel_dir: str):
        self.excel_dir = Path(excel_dir)
        self.rabona_file = self.excel_dir / 'Rabona_2026_04.xlsx'
        self.espargos_file = self.excel_dir / 'Espargos_2026_04.xlsx'
        self.settlement_file = self.excel_dir / 'Settlement_2026_04.xlsx'

    def get_next_reference(self, company: str = 'Rabona', month: int = 4) -> str:
        """
        Get next sequential reference number
        Format: YY/MM/SEQ (e.g., 26/04/1)
        """
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Expense Data']

            max_seq = 0
            for row in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=1):
                cell_value = row[0].value
                if cell_value and isinstance(cell_value, str):
                    parts = cell_value.split('/')
                    if len(parts) == 3 and parts[1] == str(month):
                        try:
                            seq = int(parts[2])
                            max_seq = max(max_seq, seq)
                        except ValueError:
                            pass

            next_seq = max_seq + 1
            return f"26/{month}/{next_seq}"

        except Exception as e:
            logger.error(f"Error getting next reference: {e}")
            return f"26/{month}/1"

    def add_expense_entry(self, company: str, entry_data: Dict) -> Dict:
        """
        Add a new expense entry to the Excel file
        entry_data should contain: reference, date, vendor, description, amount, currency,
                                  category, subcategory, payment_method, status
        """
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Expense Data']

            # Find first empty row
            empty_row = 2
            for row in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=1):
                if row[0].value is None:
                    empty_row = row[0].row
                    break

            # Map data to columns based on sheet structure
            row_num = empty_row
            sheet[f'A{row_num}'] = entry_data.get('reference', '')
            sheet[f'B{row_num}'] = entry_data.get('date', '')
            sheet[f'C{row_num}'] = entry_data.get('vendor', '')
            sheet[f'D{row_num}'] = entry_data.get('description', '')
            sheet[f'E{row_num}'] = entry_data.get('amount', 0)
            sheet[f'F{row_num}'] = entry_data.get('currency', 'EUR')
            sheet[f'G{row_num}'] = entry_data.get('category', '')
            sheet[f'H{row_num}'] = entry_data.get('subcategory', '')
            sheet[f'I{row_num}'] = entry_data.get('payment_method', '')
            sheet[f'J{row_num}'] = entry_data.get('status', 'Incomplete')

            wb.save(file)
            return {'status': 'success', 'reference': entry_data.get('reference')}

        except Exception as e:
            logger.error(f"Error adding expense: {e}")
            return {'status': 'error', 'message': str(e)}

    def get_categories(self, company: str = 'Rabona') -> List[str]:
        """Get list of available categories"""
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Settings']

            categories = []
            for row in sheet.iter_rows(min_row=2, max_row=50, min_col=1, max_col=1):
                value = row[0].value
                if value and isinstance(value, str) and value.strip():
                    categories.append(value.strip())
                elif value is None:
                    break

            return categories

        except Exception as e:
            logger.error(f"Error getting categories: {e}")
            return []

    def get_clients(self, company: str = 'Rabona') -> List[str]:
        """Get list of available clients"""
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Settings']

            clients = []
            # Clients should be in a specific range in Settings sheet
            for row in sheet.iter_rows(min_row=20, max_row=40, min_col=3, max_col=3):
                value = row[0].value
                if value and isinstance(value, str) and value.strip():
                    clients.append(value.strip())

            return clients

        except Exception as e:
            logger.error(f"Error getting clients: {e}")
            return []

    def get_payment_methods(self, company: str = 'Rabona') -> List[str]:
        """Get list of payment methods"""
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Settings']

            methods = []
            for row in sheet.iter_rows(min_row=2, max_row=20, min_col=5, max_col=5):
                value = row[0].value
                if value and isinstance(value, str) and value.strip():
                    methods.append(value.strip())

            return methods

        except Exception as e:
            logger.error(f"Error getting payment methods: {e}")
            return []

    def get_month_summary(self, company: str = 'Rabona') -> Dict:
        """Get summary statistics for current month"""
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb['Month Summary']

            summary = {}
            # Read key metrics from Month Summary sheet
            summary['total_expenses'] = sheet['B2'].value or 0
            summary['complete_entries'] = sheet['B3'].value or 0
            summary['incomplete_entries'] = sheet['B4'].value or 0

            return summary

        except Exception as e:
            logger.error(f"Error getting month summary: {e}")
            return {'total_expenses': 0, 'complete_entries': 0, 'incomplete_entries': 0}

    def mark_entry_complete(self, company: str, reference: str) -> Dict:
        """Mark an expense entry as complete"""
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Expense Data']

            for row in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=1):
                if row[0].value == reference:
                    # Status is in column J
                    sheet[f'J{row[0].row}'] = 'Complete'
                    wb.save(file)
                    return {'status': 'success'}

            return {'status': 'error', 'message': 'Reference not found'}

        except Exception as e:
            logger.error(f"Error marking entry complete: {e}")
            return {'status': 'error', 'message': str(e)}

    def get_incomplete_entries(self, company: str = 'Rabona') -> List[Dict]:
        """Get list of incomplete entries"""
        try:
            file = self.rabona_file if company == 'Rabona' else self.espargos_file
            wb = openpyxl.load_workbook(file)
            sheet = wb['Incomplete Entries']

            incomplete = []
            for row in sheet.iter_rows(min_row=2, max_row=100, min_col=1, max_col=13):
                if row[0].value:  # If reference is not empty
                    incomplete.append({
                        'reference': row[0].value,
                        'date': row[1].value,
                        'vendor': row[2].value,
                        'amount': row[3].value,
                        'missing_fields': row[4].value if len(row) > 4 else '',
                    })

            return incomplete

        except Exception as e:
            logger.error(f"Error getting incomplete entries: {e}")
            return []
