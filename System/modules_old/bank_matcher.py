"""Bank statement matching module"""

from typing import Dict, List, Optional, Tuple
from datetime import datetime, timedelta
import logging
import csv
from pathlib import Path

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class BankStatementMatcher:
    """
    Match bank statement transactions to expense entries
    Using amount + date range with user approval workflow
    """

    def __init__(self, excel_handler):
        self.excel_handler = excel_handler

    def parse_bank_statement(self, file_path: str) -> List[Dict]:
        """
        Parse bank statement CSV/Excel
        Expected columns: Date, Description, Amount, Balance
        """
        try:
            transactions = []

            if file_path.endswith('.csv'):
                transactions = self._parse_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                transactions = self._parse_excel(file_path)

            return transactions

        except Exception as e:
            logger.error(f"Error parsing bank statement: {e}")
            return []

    def _parse_csv(self, file_path: str) -> List[Dict]:
        """Parse CSV bank statement"""
        transactions = []
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        trans = {
                            'date': self._parse_date(row.get('Date', '')),
                            'description': row.get('Description', ''),
                            'amount': float(row.get('Amount', 0)),
                            'balance': float(row.get('Balance', 0)),
                        }
                        if trans['date']:
                            transactions.append(trans)
                    except (ValueError, KeyError):
                        continue

            return transactions

        except Exception as e:
            logger.error(f"Error parsing CSV: {e}")
            return []

    def _parse_excel(self, file_path: str) -> List[Dict]:
        """Parse Excel bank statement"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            transactions = []

            for row in sheet.iter_rows(min_row=2, max_row=500, min_col=1, max_col=4):
                try:
                    date = self._parse_date(str(row[0].value))
                    description = str(row[1].value) if row[1].value else ''
                    amount = float(row[2].value) if row[2].value else 0
                    balance = float(row[3].value) if row[3].value else 0

                    if date and amount != 0:
                        transactions.append({
                            'date': date,
                            'description': description,
                            'amount': amount,
                            'balance': balance,
                        })
                except (ValueError, TypeError):
                    continue

            return transactions

        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")
            return []

    def _parse_date(self, date_str: str) -> Optional[str]:
        """Parse date string to DD/MM/YYYY format"""
        if not date_str:
            return None

        formats = ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y']

        for fmt in formats:
            try:
                parsed = datetime.strptime(date_str.strip(), fmt)
                return parsed.strftime('%d/%m/%Y')
            except ValueError:
                continue

        return None

    def find_matches(
        self,
        company: str,
        bank_transactions: List[Dict],
        date_tolerance: int = 3,
        amount_tolerance: float = 0.01
    ) -> List[Dict]:
        """
        Find potential matches between bank transactions and expense entries
        Uses amount + date range matching
        Returns list of match candidates with confidence scores
        """
        matches = []

        # Get expenses from Excel
        expenses = self._get_expenses(company)

        for trans in bank_transactions:
            candidates = []

            for expense in expenses:
                # Calculate match score
                amount_diff = abs(trans['amount'] - expense['amount'])
                date_diff = self._date_diff(trans['date'], expense['date'])

                # Check if within tolerance
                if amount_diff <= amount_tolerance and date_diff <= date_tolerance:
                    confidence = self._calculate_confidence(
                        amount_diff, date_diff, trans['description'], expense['vendor']
                    )

                    candidates.append({
                        'bank_trans': trans,
                        'expense': expense,
                        'amount_diff': amount_diff,
                        'date_diff': date_diff,
                        'confidence': confidence,
                    })

            # Sort by confidence and take best matches
            if candidates:
                candidates.sort(key=lambda x: x['confidence'], reverse=True)
                matches.extend(candidates[:3])  # Top 3 candidates

        return matches

    def _get_expenses(self, company: str) -> List[Dict]:
        """Get expenses from Excel that need matching"""
        try:
            import openpyxl
            file = (
                self.excel_handler.rabona_file
                if company == 'Rabona'
                else self.excel_handler.espargos_file
            )
            wb = openpyxl.load_workbook(file, data_only=True)
            sheet = wb['Expense Data']

            expenses = []
            for row in sheet.iter_rows(min_row=2, max_row=100):
                if row[0].value:  # Reference exists
                    try:
                        expense = {
                            'reference': row[0].value,
                            'date': self._parse_date(str(row[1].value)),
                            'vendor': row[2].value,
                            'description': row[3].value,
                            'amount': float(row[4].value) if row[4].value else 0,
                            'status': row[12].value,
                        }
                        if expense['date'] and expense['amount'] > 0:
                            expenses.append(expense)
                    except (ValueError, TypeError, IndexError):
                        continue

            return expenses

        except Exception as e:
            logger.error(f"Error getting expenses: {e}")
            return []

    def _date_diff(self, date1_str: str, date2_str: str) -> int:
        """Calculate days difference between two dates (DD/MM/YYYY format)"""
        try:
            d1 = datetime.strptime(date1_str, '%d/%m/%Y')
            d2 = datetime.strptime(date2_str, '%d/%m/%Y')
            return abs((d1 - d2).days)
        except (ValueError, TypeError):
            return 999  # Large number for no match

    def _calculate_confidence(
        self,
        amount_diff: float,
        date_diff: int,
        bank_desc: str,
        vendor: str
    ) -> float:
        """
        Calculate match confidence score (0-100)
        Based on amount diff, date diff, and description similarity
        """
        # Start with 100
        score = 100

        # Penalize for amount difference
        score -= min(amount_diff * 100, 30)

        # Penalize for date difference
        score -= min(date_diff * 5, 20)

        # Bonus for description similarity
        if vendor and bank_desc:
            vendor_lower = vendor.lower()
            desc_lower = bank_desc.lower()
            if vendor_lower in desc_lower or desc_lower in vendor_lower:
                score += 10

        return max(0, score)

    def approve_match(self, company: str, reference: str, bank_trans: Dict) -> Dict:
        """
        User approves a match - link bank transaction to expense entry
        """
        try:
            result = self.excel_handler.mark_entry_complete(company, reference)
            return result

        except Exception as e:
            logger.error(f"Error approving match: {e}")
            return {'status': 'error', 'message': str(e)}

    def get_unmatched_transactions(
        self,
        company: str,
        bank_file: str
    ) -> Tuple[List[Dict], List[Dict]]:
        """
        Get unmatched bank transactions and unmatched expenses
        """
        bank_transactions = self.parse_bank_statement(bank_file)
        expenses = self._get_expenses(company)

        # Find matches
        matches = self.find_matches(company, bank_transactions)
        matched_refs = {m['expense']['reference'] for m in matches}
        matched_amounts = {m['bank_trans']['amount'] for m in matches}

        unmatched_expenses = [
            e for e in expenses
            if e['reference'] not in matched_refs
        ]

        unmatched_bank = [
            t for t in bank_transactions
            if t['amount'] not in matched_amounts
        ]

        return unmatched_expenses, unmatched_bank
